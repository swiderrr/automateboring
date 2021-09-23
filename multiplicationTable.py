#!python3
import openpyxl, sys, time
from openpyxl.styles import Font, NamedStyle 
print(sys.argv[0])
wb = openpyxl.Workbook()
ws = wb.active

bold12Font = Font(size=12, bold=True)
for k in range(1, int(sys.argv[1])+1):
    ws.cell(row=k+1, column=1).value = k
    ws.cell(row=1, column=k+1).value = k
    ws.cell(row=k+1, column=1).font = bold12Font
    ws.cell(row=1, column=k+1).font = bold12Font


for i in range(1,int(sys.argv[1])+1):
    for j in range(1,int(sys.argv[1])+1):
        ws.cell(row=i+1, column=j+1).value = i*j

wb.save(r"H:\Programowanie\Python\multiplicationTable.xlsx")

