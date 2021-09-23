#!python3

import sys, openpyxl
from openpyxl.utils.cell import get_column_letter

#if len(sys.argv) == 3:
#    try:
#        num1 = int(sys.argv[1])
#        num2 = int(sys.argv[2])
#    except Exception as e:
#        print(e)
num1 = 3
num2 = 5

wb = openpyxl.load_workbook("H:\Programowanie\Python\produceSales.xlsx")
ws = wb.active

maxColumn = ws.max_column
maxRow = ws.max_row
lastRow1 = get_column_letter(maxColumn) + str(num1)
lastRow2 = get_column_letter(maxColumn) + str(maxRow)
lastEmpty = get_column_letter(maxColumn) + str(int(num1+num2))

#przepisanie pierwszych wierszy
for row in ws['A1':lastRow1]:
    for cell in row:
        ws[cell.coordinate] = cell.value

for row in ws['A'+str(num1+1):lastRow2]:
    for cell in row:
        #print(cell.coordinate)
        #print(ws[get_column_letter(cell.column)+str(cell.row+num2)])
        #print(cell.value)
        ws[get_column_letter(cell.column)+str(cell.row+num2)] = cell.value

for row in ws['A'+str(num1+1):lastEmpty]:
    for cell in row:
        ws[cell.coordinate] = ''


        
wb.save("H:\Programowanie\Python\\testKopii.xlsx")
