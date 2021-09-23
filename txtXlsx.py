import openpyxl


def txtToxlsx(liczbaPlikow):
    wb = openpyxl.Workbook()
    ws = wb.active
    linijki = []

    for i in range(1, int(liczbaPlikow+1)):
        tekst =  open('H:\\Programowanie\\Python\\Tekst{}.txt'.format(i))
        linijki = tekst.readlines()
        for wiersz in range(0,len(linijki)):
            ws.cell(column=i, row=wiersz+1, value=linijki[wiersz])

    wb.save("H:\Programowanie\Python\\Tekst.xlsx")

def xlsxTotxt():
    wb = openpyxl.load_workbook("H:\Programowanie\Python\\Tekst.xlsx")
    ws = wb.active
    for columns in range(ws.max_column):
        tekst = open('H:\\Programowanie\\Python\\Tekst{}_{}.txt'.format(columns,columns), 'w')
        for rows in range(ws.max_row):
            if ws.cell(row=rows+1, column=columns+1).value is not None:
                tekst.write(ws.cell(row=rows+1, column=columns+1).value)
        tekst.close()

xlsxTotxt()